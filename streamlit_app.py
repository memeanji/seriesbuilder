from __future__ import annotations

import os
import json
import re
import subprocess
from datetime import datetime
from pathlib import Path

import streamlit as st


APP_DIR = Path(__file__).resolve().parent
ENV_PATH = APP_DIR / ".env"
ADS_MANAGER_URL = "https://adsmanager.facebook.com/adsmanager/manage/campaigns"
DEFAULT_ACCOUNT_ID = "1838892106940197"
DEFAULT_JOB_ROOT = Path(r"C:\meta_jobs")
DEFAULT_PROFILE_ROOT = Path(r"C:\meta_profiles")
IS_WINDOWS = os.name == "nt"
IS_STREAMLIT_CLOUD = bool(os.environ.get("STREAMLIT_RUNTIME") or os.environ.get("STREAMLIT_SHARING_MODE"))
NOTIFICATION_DEFAULTS = {
    "ENABLE_DESKTOP_ALERT": "true",
    "NOTIFY_ON_SUCCESS": "true",
    "NOTIFY_ON_ERROR": "true",
    "NOTIFY_ON_STOP": "true",
    "NOTIFY_ON_VIDEO_UPLOAD_TIMEOUT": "true",
    "NOTIFICATION_APP_NAME": "Meta Ads Automation",
    "NOTIFICATION_SUCCESS_TITLE": "작업 완료",
    "NOTIFICATION_ERROR_TITLE": "작업 중단",
    "NOTIFICATION_TIMEOUT_TITLE": "업로드 지연",
}


def env_bool(values: dict[str, str], key: str, default: str = "false") -> bool:
    return values.get(key, default).strip().lower() in {"1", "true", "yes", "y", "on"}


def read_env(path: Path = ENV_PATH) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip()
    return values


def write_env(values: dict[str, str], path: Path = ENV_PATH) -> str:
    mode = values.get("CAMPAIGN_MODE", "IMAGE_ONLY")
    lines = [
        f"AD_ACCOUNT_ID={values.get('AD_ACCOUNT_ID', '')}",
        f"CAMPAIGN_NAME={values.get('CAMPAIGN_NAME', '')}",
        "",
        f"CAMPAIGN_MODE={mode}",
        f"DRY_RUN={values.get('DRY_RUN', 'true')}",
        "",
        f"ADSET_START_INDEX={values.get('ADSET_START_INDEX', '1')}",
        f"ADSET_COUNT={values.get('ADSET_COUNT', '1')}",
        f"ADSET_DAILY_BUDGET={values.get('ADSET_DAILY_BUDGET', '100000')}",
        f"CDP_URL={values.get('CDP_URL', 'http://127.0.0.1:9222')}",
        f"SCHEDULE_TIME={values.get('SCHEDULE_TIME', '05:00')}",
        f"LANDING_PATH_NUMBER={values.get('LANDING_PATH_NUMBER', '100')}",
        "",
        f"ENABLE_DESKTOP_ALERT={values.get('ENABLE_DESKTOP_ALERT', NOTIFICATION_DEFAULTS['ENABLE_DESKTOP_ALERT'])}",
        f"NOTIFY_ON_SUCCESS={values.get('NOTIFY_ON_SUCCESS', NOTIFICATION_DEFAULTS['NOTIFY_ON_SUCCESS'])}",
        f"NOTIFY_ON_ERROR={values.get('NOTIFY_ON_ERROR', NOTIFICATION_DEFAULTS['NOTIFY_ON_ERROR'])}",
        f"NOTIFY_ON_STOP={values.get('NOTIFY_ON_STOP', NOTIFICATION_DEFAULTS['NOTIFY_ON_STOP'])}",
        f"NOTIFY_ON_VIDEO_UPLOAD_TIMEOUT={values.get('NOTIFY_ON_VIDEO_UPLOAD_TIMEOUT', NOTIFICATION_DEFAULTS['NOTIFY_ON_VIDEO_UPLOAD_TIMEOUT'])}",
        f"NOTIFICATION_APP_NAME={values.get('NOTIFICATION_APP_NAME', NOTIFICATION_DEFAULTS['NOTIFICATION_APP_NAME'])}",
        f"NOTIFICATION_SUCCESS_TITLE={values.get('NOTIFICATION_SUCCESS_TITLE', NOTIFICATION_DEFAULTS['NOTIFICATION_SUCCESS_TITLE'])}",
        f"NOTIFICATION_ERROR_TITLE={values.get('NOTIFICATION_ERROR_TITLE', NOTIFICATION_DEFAULTS['NOTIFICATION_ERROR_TITLE'])}",
        f"NOTIFICATION_TIMEOUT_TITLE={values.get('NOTIFICATION_TIMEOUT_TITLE', NOTIFICATION_DEFAULTS['NOTIFICATION_TIMEOUT_TITLE'])}",
        "",
        f"WAIT_BASE_RETRY_COUNT={values.get('WAIT_BASE_RETRY_COUNT', '5')}",
        f"WAIT_BASE_RETRY_INTERVAL_MS={values.get('WAIT_BASE_RETRY_INTERVAL_MS', '1500')}",
        f"WAIT_EXTENDED_RETRY_COUNT={values.get('WAIT_EXTENDED_RETRY_COUNT', '5')}",
        f"WAIT_EXTENDED_RETRY_INTERVAL_MS={values.get('WAIT_EXTENDED_RETRY_INTERVAL_MS', '7000')}",
        f"VIDEO_UPLOAD_TIMEOUT_MS={values.get('VIDEO_UPLOAD_TIMEOUT_MS', '120000')}",
        f"VIDEO_UPLOAD_FALLBACK_WAIT_MS={values.get('VIDEO_UPLOAD_FALLBACK_WAIT_MS', '60000')}",
        f"MODE_01_WAIT_MS={values.get('MODE_01_WAIT_MS', '7000')}",
        f"MODE_02_WAIT_MS={values.get('MODE_02_WAIT_MS', '7000')}",
        f"MODE_03_WAIT_MS={values.get('MODE_03_WAIT_MS', '9000')}",
        f"MODE_04_WAIT_MS={values.get('MODE_04_WAIT_MS', '8000')}",
        "",
    ]

    if mode == "BLOG_MIXED":
        blog_actual_creatives = int(values.get("AD_CREATIVE_COUNT", "4") or "4") + 1
        blog_image_creatives = max(blog_actual_creatives - 1, 1)
        lines.extend(
            [
                f"AD_CREATIVE_COUNT={values.get('AD_CREATIVE_COUNT', '4')}",
                f"BLOG_IMAGE_ADS_PER_ADSET={blog_image_creatives}",
                "BLOG_VIDEO_ADS_PER_ADSET=1",
                f"BLOG_TOTAL_ADS_PER_ADSET={blog_actual_creatives}",
                "BLOG_ADSET_NAME_PREFIX=f_i_b_o_l",
                "BLOG_IMAGE_AD_NAME_PREFIX=f_i_b_o_l",
                "BLOG_VIDEO_AD_NAME_PREFIX=f_v_b_o_l",
                "DATE_FORMAT=MMDD",
                "",
                f"BLOG_ASSET_ROOT={values.get('BLOG_ASSET_ROOT', '')}",
                "",
            ]
        )
        adset_count = int(values.get("ADSET_COUNT", "1") or "1")
        for index in range(1, adset_count + 1):
            lines.append(f"BLOG_LANDING_URL_{index}={values.get(f'BLOG_LANDING_URL_{index}', '')}")
        lines.append("")
    elif mode == "VIDEO_ONLY":
        lines.extend(
            [
                f"AD_CREATIVE_COUNT={values.get('AD_CREATIVE_COUNT', '1')}",
                "AD_FORMAT=video",
                "VIDEO_ONLY_AD_NAME_PREFIX=f_v_o_l",
                "DATE_FORMAT=MMDD",
                f"VIDEO_ONLY_ASSET_ROOT={values.get('VIDEO_ONLY_ASSET_ROOT', '')}",
                "",
            ]
        )
    elif mode in {"VIDEO_ONLY_CBO", "IMAGE_ONLY_CBO"}:
        is_image_cbo = mode == "IMAGE_ONLY_CBO"
        mode_prefix = "IMAGE_ONLY_CBO" if is_image_cbo else "VIDEO_ONLY_CBO"
        folder_key = "IMAGE_ONLY_CBO_IMAGE_FOLDER" if is_image_cbo else "VIDEO_ONLY_CBO_VIDEO_FOLDER"
        ad_prefix_key = "IMAGE_ONLY_CBO_AD_NAME_PREFIX" if is_image_cbo else "VIDEO_ONLY_CBO_AD_NAME_PREFIX"
        ad_prefix_value = "f_i_b_o_l" if is_image_cbo else "f_v_b_o_l"
        lines.extend(
            [
                f"AD_CREATIVE_COUNT={values.get('AD_CREATIVE_COUNT', '0')}",
                f"AD_FORMAT={'image' if is_image_cbo else 'video'}",
                f"{ad_prefix_key}={ad_prefix_value}",
                "DATE_FORMAT=MMDD",
                f"CAMPAIGN_BUDGET={values.get('CAMPAIGN_BUDGET', '')}",
                f"{folder_key}={values.get(folder_key, '')}",
                f"{mode_prefix}_AUTO_LANDING_URL=true",
                "",
            ]
        )
        adset_count = int(values.get("ADSET_COUNT", "0") or "0") + 1
        creative_count = int(values.get("AD_CREATIVE_COUNT", "0") or "0") + 1
        for index in range(1, adset_count + 1):
            key = f"{mode_prefix}_ADSET_NAME_{index}"
            lines.append(f"{key}={values.get(key, '')}")
        for index in range(1, adset_count * creative_count + 1):
            key = f"{mode_prefix}_AD_NAME_{index}"
            lines.append(f"{key}={values.get(key, '')}")
        lines.append("")
    else:
        lines.extend(
            [
                f"AD_CREATIVE_COUNT={values.get('AD_CREATIVE_COUNT', '4')}",
                "AD_FORMAT=image",
                f"IMAGE_ONLY_UPLOAD_MODE={values.get('IMAGE_ONLY_UPLOAD_MODE', '')}",
                f"IMAGE_ONLY_ASSET_ROOT={values.get('IMAGE_ONLY_ASSET_ROOT', '')}",
                f"MEDIA_FOLDER_PATH={values.get('MEDIA_FOLDER_PATH', '')}",
                "",
            ]
        )

    lines.extend(
        [
            "QUICK_TEST_CREATIVE_STEP=false",
            f"QUICK_TEST_AD_NAME={values.get('QUICK_TEST_AD_NAME', 'f_i_o_l_0518_1')}",
            "",
        ]
    )
    content = "\n".join(lines)
    path.write_text(content, encoding="utf-8")
    return content


def run_command(command: list[str], timeout: int = 300) -> tuple[int, str]:
    completed = subprocess.run(
        command,
        cwd=APP_DIR,
        text=True,
        capture_output=True,
        timeout=timeout,
        shell=False,
        encoding="utf-8",
        errors="replace",
    )
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part)
    return completed.returncode, output


def start_powershell(command: str) -> None:
    if not IS_WINDOWS:
        raise RuntimeError("PowerShell automation launcher is available only on the local Windows PC.")
    subprocess.Popen(
        [
            "powershell.exe",
            "-NoExit",
            "-ExecutionPolicy",
            "Bypass",
            "-Command",
            command,
        ],
        cwd=APP_DIR,
        creationflags=subprocess.CREATE_NEW_CONSOLE,
    )


def list_job_dirs(job_root: Path) -> list[Path]:
    if not job_root.exists():
        return []
    return sorted([path for path in job_root.iterdir() if path.is_dir() and (path / "mapping.xlsx").exists()])


def read_mapping_xlsx(mapping_path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("mapping.xlsx 미리보기를 위해 openpyxl이 필요합니다. `pip install -r requirements.txt`를 먼저 실행해주세요.") from exc

    workbook = load_workbook(mapping_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value or "").strip() for value in rows[0]]
    records: list[dict[str, str]] = []
    for row in rows[1:]:
        record = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row[index] if index < len(row) else ""
            record[header] = "" if value is None else str(value).strip()
        if any(record.values()):
            records.append(record)
    return records


def write_job_mapping_cache(job_dir: Path, records: list[dict[str, str]]) -> Path:
    cache_path = job_dir / "mapping.cache.json"
    cache_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    return cache_path


def read_job_state(job_dir: Path) -> dict:
    state_path = job_dir / "job_state.json"
    if not state_path.exists():
        return {}
    try:
        return json.loads(state_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"status": "invalid_state_file", "error": "job_state.json parse failed"}


def default_blog_root() -> str:
    mmdd = datetime.now().strftime("%m%d")
    return str(Path.home() / "Desktop" / f"F_I_B_O_L_{mmdd}")


def default_image_root() -> str:
    mmdd = datetime.now().strftime("%m%d")
    return str(Path.home() / "Desktop" / f"F_I_O_L_{mmdd}")


def default_video_root() -> str:
    yymmdd = datetime.now().strftime("%y%m%d")
    return str(Path.home() / "Desktop" / f"{yymmdd} 올레놀샷 틱톡세팅")


def expected_blog_folder_name(adset_index: int, budget: str, schedule_time: str, image_count: int = 4) -> str:
    mmdd = datetime.now().strftime("%m%d")
    budget_manwon = int(int(budget or "0") / 10000)
    hour = schedule_time.split(":", 1)[0].zfill(2)
    return f"{mmdd} {adset_index}번 광고세트-일예산 {budget_manwon}만원-이미지 {image_count}개 + 영상 1개 익일 {hour}시"


def build_blog_adset_name(index: int) -> str:
    mmdd = datetime.now().strftime("%m%d")
    return f"f_i_b_o_l_{mmdd}_{index}"


def build_blog_image_ad_name(index: int) -> str:
    mmdd = datetime.now().strftime("%m%d")
    return f"f_i_b_o_l_{mmdd}_{index}"


def build_blog_video_ad_name(index: int) -> str:
    mmdd = datetime.now().strftime("%m%d")
    return f"f_v_b_o_l_{mmdd}_{index}"


def expected_image_folder_name(adset_index: int, budget: str, creative_count: int, schedule_time: str) -> str:
    budget_manwon = int(int(budget or "0") / 10000)
    hour = schedule_time.split(":", 1)[0].zfill(2)
    return f"메타 리타겟 소재-{adset_index}번 세트-일예산 {budget_manwon}만원_익일 {hour}시 세팅"


def build_image_only_ad_name(index: int) -> str:
    mmdd = datetime.now().strftime("%m%d")
    return f"f_i_o_l_{mmdd}_{index}"


def expected_video_folder_name() -> str:
    yymmdd = datetime.now().strftime("%y%m%d")
    return f"{yymmdd} 올레놀샷 틱톡세팅"


def expected_video_child_folder_name() -> str:
    mmdd = datetime.now().strftime("%m%d")
    return f"{mmdd} 올레놀샷 CBO 캠페인-1"


def format_budget_preview(value: str) -> str:
    raw = str(value or "").replace(",", "").strip()
    if not raw.isdigit() or int(raw) <= 0:
        return ""
    return f"{int(raw):,}"


def build_video_only_cbo_ad_name(index: int) -> str:
    mmdd = datetime.now().strftime("%m%d")
    return f"f_v_b_o_l_{mmdd}_{index}"


def build_image_only_cbo_ad_name(index: int) -> str:
    mmdd = datetime.now().strftime("%m%d")
    return f"f_i_b_o_l_{mmdd}_{index}"


def build_video_only_cbo_adset_name(index: int) -> str:
    mmdd = datetime.now().strftime("%m%d")
    return f"{mmdd} CBO 광고세트 -{index}"


def build_image_only_cbo_adset_name(index: int) -> str:
    mmdd = datetime.now().strftime("%m%d")
    return f"{mmdd} CBO 광고세트 -{index}"


def default_landing_url(ad_name: str, path_number: str = "100") -> str:
    path = str(path_number or "100").strip()
    if not path.isdigit():
        path = "100"
    return f"https://repurely.com/surl/P/{path}?utm_source=f&utm_medium=f&utm_campaign={ad_name}"

def find_video_file_for_preview(ad_name: str, folder: str) -> tuple[str, list[str]]:
    expected = [f"{ad_name}.mp4", f"{ad_name}.mov", f"{ad_name}.m4v"]
    folder_path = Path(folder).expanduser()
    for filename in expected:
        candidate = folder_path / filename
        if candidate.exists():
            return str(candidate), expected
    return "", expected


def find_image_file_for_preview(ad_name: str, folder: str) -> tuple[str, list[str]]:
    expected = [f"{ad_name}.png", f"{ad_name}.jpg", f"{ad_name}.jpeg", f"{ad_name}.webp", f"{ad_name}.gif"]
    folder_path = Path(folder).expanduser()
    for filename in expected:
        candidate = folder_path / filename
        if candidate.exists():
            return str(candidate), expected
    return "", expected


def list_video_stems_for_preview(folder: str) -> list[str]:
    folder_path = Path(folder).expanduser()
    if not folder_path.exists() or not folder_path.is_dir():
        return []
    files = [
        item
        for item in folder_path.iterdir()
        if item.is_file() and item.suffix.lower() in [".mp4", ".mov", ".m4v"]
    ]
    return [item.stem for item in sorted(files, key=lambda item: item.name.lower())]


def list_image_stems_for_preview(folder: str) -> list[str]:
    folder_path = Path(folder).expanduser()
    if not folder_path.exists() or not folder_path.is_dir():
        return []
    files = [
        item
        for item in folder_path.iterdir()
        if item.is_file() and item.suffix.lower() in [".png", ".jpg", ".jpeg", ".webp", ".gif"]
    ]
    return [item.stem for item in sorted(files, key=lambda item: item.name.lower())]


def read_child_folder_names(root: str, limit: int) -> list[str]:
    root_path = Path(root).expanduser()
    if not root_path.exists() or not root_path.is_dir():
        return []

    folders = [item.name for item in root_path.iterdir() if item.is_dir()]
    return sorted(folders, key=folder_sort_key)[:limit]


def folder_sort_key(folder_name: str) -> tuple[int, str]:
    patterns = [
        r"메타\s*리타겟\s*소재\s*-\s*(\d+)\s*번\s*세트",
        r"(\d+)\s*번\s*세트",
        r"(\d+)\s*번\s*광고세트",
        r"adset[_\-\s]*(\d+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, folder_name, flags=re.IGNORECASE)
        if match:
            return int(match.group(1)), folder_name
    return 999999, folder_name


def validate_form(values: dict[str, str]) -> list[str]:
    errors: list[str] = []
    if not values.get("CAMPAIGN_NAME", "").strip():
        errors.append("CAMPAIGN_NAME is required.")
    if not values.get("AD_ACCOUNT_ID", "").strip():
        errors.append("AD_ACCOUNT_ID is required.")
    if values.get("CAMPAIGN_MODE") == "BLOG_MIXED":
        if not values.get("BLOG_ASSET_ROOT", "").strip():
            errors.append("BLOG_ASSET_ROOT is required for BLOG_MIXED.")
        adset_count = int(values.get("ADSET_COUNT", "1") or "1")
        for index in range(1, adset_count + 1):
            if not values.get(f"BLOG_LANDING_URL_{index}", "").strip():
                errors.append(f"BLOG_LANDING_URL_{index} is required.")
    elif values.get("CAMPAIGN_MODE") == "IMAGE_ONLY":
        if not str(values.get("LANDING_PATH_NUMBER", "100")).strip().isdigit():
            errors.append("LANDING_PATH_NUMBER must be a positive number.")
        if values.get("IMAGE_ONLY_UPLOAD_MODE") != "LEGACY" and not (values.get("IMAGE_ONLY_ASSET_ROOT", "").strip() or values.get("MEDIA_FOLDER_PATH", "").strip()):
            errors.append("IMAGE_ONLY per-ad upload requires IMAGE_ONLY_ASSET_ROOT or MEDIA_FOLDER_PATH.")
    elif values.get("CAMPAIGN_MODE") == "VIDEO_ONLY":
        if not values.get("VIDEO_ONLY_ASSET_ROOT", "").strip():
            errors.append("VIDEO_ONLY_ASSET_ROOT is required for VIDEO_ONLY.")
    elif values.get("CAMPAIGN_MODE") in {"VIDEO_ONLY_CBO", "IMAGE_ONLY_CBO"}:
        mode = values.get("CAMPAIGN_MODE")
        is_image_cbo = mode == "IMAGE_ONLY_CBO"
        mode_prefix = "IMAGE_ONLY_CBO" if is_image_cbo else "VIDEO_ONLY_CBO"
        folder_key = "IMAGE_ONLY_CBO_IMAGE_FOLDER" if is_image_cbo else "VIDEO_ONLY_CBO_VIDEO_FOLDER"
        folder_label = "Image folder" if is_image_cbo else "Video folder"
        budget_preview = format_budget_preview(values.get("CAMPAIGN_BUDGET", ""))
        if not budget_preview:
            errors.append(f"Campaign budget must be a positive integer for {mode}.")
        if not str(values.get("LANDING_PATH_NUMBER", "100")).strip().isdigit():
            errors.append("LANDING_PATH_NUMBER must be a positive number.")
        media_folder = values.get(folder_key, "").strip()
        if not media_folder:
            errors.append(f"{folder_label} is required for {mode}.")
        elif not Path(media_folder).expanduser().exists():
            errors.append(f"{folder_label} does not exist: {media_folder}")

        adset_total = 1
        creative_total = int(values.get("AD_CREATIVE_COUNT", "0") or "0") + 1
        names: list[str] = []
        for index in range(1, adset_total * creative_total + 1):
            ad_name = values.get(f"{mode_prefix}_AD_NAME_{index}", "").strip() or (build_image_only_cbo_ad_name(index) if is_image_cbo else build_video_only_cbo_ad_name(index))
            if ad_name in names:
                errors.append(f"Duplicate ad name: {ad_name}")
            names.append(ad_name)
            media_file, expected = (find_image_file_for_preview(ad_name, media_folder) if is_image_cbo else find_video_file_for_preview(ad_name, media_folder))
            if not media_file:
                media_type = "Image" if is_image_cbo else "Video"
                errors.append(f"{media_type} file not found for ad name: {ad_name}. Expected one of: {', '.join(expected)}")
    return errors


def show_command_result(title: str, command: list[str], code: int, output: str) -> None:
    if code == 0:
        st.success(f"{title} completed")
    else:
        st.error(f"{title} failed: exit code {code}")
    st.caption("Command")
    st.code(" ".join(command), language="text")
    st.caption("Output")
    st.code(output if output.strip() else "(empty output)", language="text")


st.set_page_config(page_title="Meta Ads Automation", layout="wide")

title_col, guide_col = st.columns([0.62, 0.38], vertical_alignment="center")
with title_col:
    st.title("Meta Ads Automation")
with guide_col:
    st.caption(
        "회사 공유 사용 안내: Streamlit Cloud는 설정/프리뷰용이고, 실제 Meta 자동화는 각자 Windows PC에서 로컬 실행해야 합니다."
    )

with st.expander("회사 사람들과 같이 쓰는 방법"):
    st.markdown(
        """
        - Streamlit Cloud 앱은 입력값 확인, preview, URL 예시 확인용입니다.
        - 실제 자동화는 각 사용자 PC의 Chrome 로그인 세션, 로컬 소재 폴더, PowerShell을 사용하므로 각자 Windows PC에서 실행해야 합니다.
        - 처음 쓰는 사람은 GitHub 저장소를 받은 뒤 `npm install`, `pip install -r requirements.txt`, `streamlit run streamlit_app.py` 순서로 실행합니다.
        - 각 사용자는 자기 PC에서 `.env`를 따로 저장하고, 자기 Meta 계정으로 Chrome CDP를 열어 자동화를 실행합니다.
        - 업데이트가 필요할 때는 로컬 폴더에서 `git pull origin main` 후 다시 Streamlit을 실행하면 됩니다.
        """
    )

env = read_env()

with st.expander("Local job runner MVP: C:\\meta_jobs 기반 실행", expanded=False):
    job_root = Path(st.text_input("Job root", value=env.get("LOCAL_JOB_ROOT", str(DEFAULT_JOB_ROOT))))
    profile_root = Path(st.text_input("Profile root", value=env.get("LOCAL_PROFILE_ROOT", str(DEFAULT_PROFILE_ROOT))))
    profile_id = st.selectbox("Browser profile", ["profile_01", "profile_02", "profile_03"], index=0)
    account_id_for_job = st.text_input("Job runner ad account ID", value=env.get("AD_ACCOUNT_ID", DEFAULT_ACCOUNT_ID), key="job_runner_account_id")

    job_dirs = list_job_dirs(job_root)
    if not job_dirs:
        st.info("job root 아래에 `mapping.xlsx`가 들어있는 job 폴더가 아직 없습니다. 예: C:\\meta_jobs\\job_01\\mapping.xlsx")
    else:
        selected_job_name = st.selectbox("Job folder", [path.name for path in job_dirs])
        selected_job = next(path for path in job_dirs if path.name == selected_job_name)
        mapping_path = selected_job / "mapping.xlsx"
        state = read_job_state(selected_job)

        try:
            mapping_records = read_mapping_xlsx(mapping_path)
            st.write(f"Mapping rows: `{len(mapping_records)}`")
            st.dataframe(mapping_records, use_container_width=True, hide_index=True)
        except Exception as exc:
            mapping_records = []
            st.error(str(exc))

        state_cols = st.columns(4)
        state_cols[0].metric("Status", state.get("status", "not_started"))
        state_cols[1].metric("Last completed", state.get("last_completed_item", "-"))
        state_cols[2].metric("Failed item", state.get("failed_item", "-"))
        state_cols[3].metric("Failed step", state.get("failed_step", "-"))
        if state.get("error"):
            st.error(state["error"])

        runner_env = {
            "LOCAL_JOB_DIR": str(selected_job),
            "LOCAL_JOB_ID": selected_job.name,
            "LOCAL_PROFILE_ROOT": str(profile_root),
            "LOCAL_PROFILE_ID": profile_id,
            "AD_ACCOUNT_ID": account_id_for_job,
            "ADS_MANAGER_URL": ADS_MANAGER_URL,
            "SLOW_MO_MS": "120",
            "JOB_RUNNER_DRY_RUN": "false",
        }
        if mapping_records:
            write_job_mapping_cache(selected_job, mapping_records)

        runner_command = (
            "$env:LOCAL_JOB_DIR='{LOCAL_JOB_DIR}'; "
            "$env:LOCAL_JOB_ID='{LOCAL_JOB_ID}'; "
            "$env:LOCAL_PROFILE_ROOT='{LOCAL_PROFILE_ROOT}'; "
            "$env:LOCAL_PROFILE_ID='{LOCAL_PROFILE_ID}'; "
            "$env:AD_ACCOUNT_ID='{AD_ACCOUNT_ID}'; "
            "$env:ADS_MANAGER_URL='{ADS_MANAGER_URL}'; "
            "$env:SLOW_MO_MS='{SLOW_MO_MS}'; "
            "$env:JOB_RUNNER_DRY_RUN='{JOB_RUNNER_DRY_RUN}'; "
            "node src/job-runner.js"
        ).format(**runner_env)

        job_col1, job_col2, job_col3 = st.columns(3)
        with job_col1:
            if st.button("Run selected job", disabled=not IS_WINDOWS or not mapping_records):
                try:
                    start_powershell(runner_command)
                    st.success("Local job runner terminal opened.")
                except Exception as exc:
                    st.error(f"Local job runner cannot start here: {exc}")
        with job_col2:
            if st.button("Restart from failed item", disabled=not IS_WINDOWS or not mapping_records):
                try:
                    start_powershell(runner_command)
                    st.success("Restart terminal opened. job_state.json 기준으로 실패 소재부터 재개합니다.")
                except Exception as exc:
                    st.error(f"Local job runner cannot restart here: {exc}")
        with job_col3:
            st.caption("Stop 기능은 다음 단계에서 큐/프로세스 관리와 함께 붙입니다.")

        log_path = selected_job / "logs" / "run.log"
        error_log_path = selected_job / "logs" / "error.log"
        if log_path.exists():
            with st.expander("logs/run.log", expanded=False):
                st.code(log_path.read_text(encoding="utf-8", errors="replace")[-8000:], language="text")
        if error_log_path.exists():
            with st.expander("logs/error.log", expanded=False):
                st.code(error_log_path.read_text(encoding="utf-8", errors="replace")[-8000:], language="text")

with st.expander("공통 wait/retry 설정", expanded=False):
    wait_base_retry_count = st.number_input("기본 재시도 횟수", min_value=1, max_value=20, value=int(env.get("WAIT_BASE_RETRY_COUNT", "5") or "5"))
    wait_base_retry_interval = st.number_input("기본 재시도 간격(ms)", min_value=500, max_value=10000, value=int(env.get("WAIT_BASE_RETRY_INTERVAL_MS", "1500") or "1500"), step=500)
    wait_extended_retry_count = st.number_input("확장 재시도 횟수", min_value=1, max_value=20, value=int(env.get("WAIT_EXTENDED_RETRY_COUNT", "5") or "5"))
    wait_extended_retry_interval = st.number_input("확장 재시도 간격(ms)", min_value=1000, max_value=30000, value=int(env.get("WAIT_EXTENDED_RETRY_INTERVAL_MS", "7000") or "7000"), step=1000)
    video_upload_timeout_ms = st.number_input("영상 업로드 확인 timeout(ms)", min_value=30000, max_value=300000, value=int(env.get("VIDEO_UPLOAD_TIMEOUT_MS", "120000") or "120000"), step=10000)
    video_upload_fallback_wait_ms = st.number_input("영상 fallback 대기(ms)", min_value=30000, max_value=90000, value=int(env.get("VIDEO_UPLOAD_FALLBACK_WAIT_MS", "60000") or "60000"), step=10000)
    mode_01_wait_ms = st.number_input("mode_01_wait BLOG_MIXED(ms)", min_value=1000, max_value=30000, value=int(env.get("MODE_01_WAIT_MS", "7000") or "7000"), step=1000)
    mode_02_wait_ms = st.number_input("mode_02_wait IMAGE_ONLY(ms)", min_value=1000, max_value=30000, value=int(env.get("MODE_02_WAIT_MS", "7000") or "7000"), step=1000)
    mode_03_wait_ms = st.number_input("mode_03_wait VIDEO_ONLY_CBO(ms)", min_value=1000, max_value=30000, value=int(env.get("MODE_03_WAIT_MS", "9000") or "9000"), step=1000)
    mode_04_wait_ms = st.number_input("mode_04_wait IMAGE_ONLY_CBO(ms)", min_value=1000, max_value=30000, value=int(env.get("MODE_04_WAIT_MS", "8000") or "8000"), step=1000)

with st.sidebar:
    st.subheader("Actions")
    if not IS_WINDOWS:
        st.info("Cloud mode: Windows-only local automation buttons are disabled. Use this page for settings and preview, then run automation on your local PC.")
    if st.button("Pull latest code"):
        command = ["git", "pull", "origin", "main"]
        code, output = run_command(command)
        show_command_result("git pull", command, code, output)

    if st.button("Install npm packages", disabled=not IS_WINDOWS):
        command = ["npm.cmd" if IS_WINDOWS else "npm", "install"]
        code, output = run_command(command, timeout=600)
        show_command_result("npm install", command, code, output)

    if st.button("Open .env in VS Code", disabled=not IS_WINDOWS):
        subprocess.Popen(["code.cmd", str(ENV_PATH)], cwd=APP_DIR, shell=False)
        st.info(".env open requested.")

    if st.button("Open Chrome CDP", disabled=not IS_WINDOWS):
        account_id = env.get("AD_ACCOUNT_ID", DEFAULT_ACCOUNT_ID)
        url = f"{ADS_MANAGER_URL}?act={account_id}"
        chrome = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
        command = f'& "{chrome}" --remote-debugging-port=9222 --user-data-dir="C:\\chrome-debug" "{url}"'
        try:
            start_powershell(command)
            st.info("Chrome CDP PowerShell opened.")
        except Exception as exc:
            st.error(f"Chrome CDP launcher is unavailable here: {exc}")

    if st.button("Open automation terminal", disabled=not IS_WINDOWS):
        try:
            start_powershell("npm run open-campaign")
            st.info("Automation PowerShell opened.")
        except Exception as exc:
            st.error(f"Automation terminal launcher is unavailable here: {exc}")

    st.subheader("Notifications")
    enable_desktop_alert = st.toggle(
        "데스크톱 알림 사용",
        value=env_bool(env, "ENABLE_DESKTOP_ALERT", NOTIFICATION_DEFAULTS["ENABLE_DESKTOP_ALERT"]),
    )
    notify_on_success = st.toggle(
        "작업 완료 시 알림",
        value=env_bool(env, "NOTIFY_ON_SUCCESS", NOTIFICATION_DEFAULTS["NOTIFY_ON_SUCCESS"]),
    )
    notify_on_error = st.toggle(
        "에러 발생 시 알림",
        value=env_bool(env, "NOTIFY_ON_ERROR", NOTIFICATION_DEFAULTS["NOTIFY_ON_ERROR"]),
    )
    notify_on_stop = st.toggle(
        "검증/중단 시 알림",
        value=env_bool(env, "NOTIFY_ON_STOP", NOTIFICATION_DEFAULTS["NOTIFY_ON_STOP"]),
    )
    notify_on_video_upload_timeout = st.toggle(
        "영상 업로드 타임아웃 알림",
        value=env_bool(env, "NOTIFY_ON_VIDEO_UPLOAD_TIMEOUT", NOTIFICATION_DEFAULTS["NOTIFY_ON_VIDEO_UPLOAD_TIMEOUT"]),
    )

    st.divider()
    st.markdown(
        """
        <div style="font-size:0.82rem; line-height:1.45; color:rgba(49,51,63,0.70);">
          <div style="margin-bottom:5.2px;">Meta Ads Automation</div>
          <a href="https://repurely.com/" target="_blank" style="display:block; margin-bottom:5.2px; color:inherit; text-decoration:none;">https://repurely.com</a>
          <div>@memeanji</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


st.subheader("Campaign Settings")

left, right = st.columns(2)
with left:
    campaign_mode_options = ["BLOG_MIXED", "IMAGE_ONLY", "VIDEO_ONLY_CBO", "IMAGE_ONLY_CBO"]
    current_campaign_mode = env.get("CAMPAIGN_MODE", "IMAGE_ONLY")
    if current_campaign_mode not in campaign_mode_options:
        current_campaign_mode = "IMAGE_ONLY"
    campaign_mode = st.selectbox(
        "Campaign mode",
        campaign_mode_options,
        index=campaign_mode_options.index(current_campaign_mode),
    )
    dry_run = st.toggle("DRY_RUN", value=env.get("DRY_RUN", "true").lower() == "true")
    ad_account_id = st.text_input("Ad account ID", value=env.get("AD_ACCOUNT_ID", DEFAULT_ACCOUNT_ID))
    campaign_name = st.text_input("Campaign name", value=env.get("CAMPAIGN_NAME", ""))

with right:
    if campaign_mode in {"VIDEO_ONLY_CBO", "IMAGE_ONLY_CBO"}:
        adset_count = 0
        st.info("CBO campaign structure: 1 campaign / 1 adset. Choose only the ad count below.")
    else:
        adset_count_min = 1 if campaign_mode == "BLOG_MIXED" else 0
        adset_count_default = max(adset_count_min, int(env.get("ADSET_COUNT", "3") or "3"))
        adset_count_label = "Adset count (BLOG_MIXED actual adsets)" if campaign_mode == "BLOG_MIXED" else "Adset count"
        if campaign_mode == "BLOG_MIXED":
            adset_count_help = "입력한 숫자 그대로 실제 광고세트 개수입니다. 광고세트 안의 소재 수는 BLOG_MIXED의 Ad creative count에서 정하고, 마지막 소재 1개는 영상입니다."
        else:
            adset_count_help = "기존 IMAGE_ONLY/VIDEO_ONLY 흐름에서는 입력값이 복제 기준으로 저장되어 실제 광고세트는 입력값 + 1개로 구성됩니다. 예: 0 입력 -> 1개, 2 입력 -> 3개."
        adset_count = st.number_input(
            adset_count_label,
            min_value=adset_count_min,
            max_value=100,
            value=adset_count_default,
            help=adset_count_help,
        )
    daily_budget = st.text_input("Daily budget", value=env.get("ADSET_DAILY_BUDGET", "300000"))
    schedule_time = st.text_input("Schedule time", value=env.get("SCHEDULE_TIME", "05:00"))
    cdp_url = st.text_input("CDP URL", value=env.get("CDP_URL", "http://127.0.0.1:9222"))

next_env: dict[str, str] = {
    "AD_ACCOUNT_ID": ad_account_id,
    "CAMPAIGN_NAME": campaign_name,
    "CAMPAIGN_MODE": campaign_mode,
    "DRY_RUN": str(dry_run).lower(),
    "ADSET_START_INDEX": "1",
    "ADSET_COUNT": str(adset_count),
    "ADSET_DAILY_BUDGET": daily_budget,
    "CDP_URL": cdp_url,
    "SCHEDULE_TIME": schedule_time,
    "LANDING_PATH_NUMBER": env.get("LANDING_PATH_NUMBER", env.get("REPURELY_PATH_NUMBER", "100")),
    "ENABLE_DESKTOP_ALERT": str(enable_desktop_alert).lower(),
    "NOTIFY_ON_SUCCESS": str(notify_on_success).lower(),
    "NOTIFY_ON_ERROR": str(notify_on_error).lower(),
    "NOTIFY_ON_STOP": str(notify_on_stop).lower(),
    "NOTIFY_ON_VIDEO_UPLOAD_TIMEOUT": str(notify_on_video_upload_timeout).lower(),
    "NOTIFICATION_APP_NAME": env.get("NOTIFICATION_APP_NAME", NOTIFICATION_DEFAULTS["NOTIFICATION_APP_NAME"]),
    "NOTIFICATION_SUCCESS_TITLE": env.get("NOTIFICATION_SUCCESS_TITLE", NOTIFICATION_DEFAULTS["NOTIFICATION_SUCCESS_TITLE"]),
    "NOTIFICATION_ERROR_TITLE": env.get("NOTIFICATION_ERROR_TITLE", NOTIFICATION_DEFAULTS["NOTIFICATION_ERROR_TITLE"]),
    "NOTIFICATION_TIMEOUT_TITLE": env.get("NOTIFICATION_TIMEOUT_TITLE", NOTIFICATION_DEFAULTS["NOTIFICATION_TIMEOUT_TITLE"]),
    "WAIT_BASE_RETRY_COUNT": str(wait_base_retry_count),
    "WAIT_BASE_RETRY_INTERVAL_MS": str(wait_base_retry_interval),
    "WAIT_EXTENDED_RETRY_COUNT": str(wait_extended_retry_count),
    "WAIT_EXTENDED_RETRY_INTERVAL_MS": str(wait_extended_retry_interval),
    "VIDEO_UPLOAD_TIMEOUT_MS": str(video_upload_timeout_ms),
    "VIDEO_UPLOAD_FALLBACK_WAIT_MS": str(video_upload_fallback_wait_ms),
    "MODE_01_WAIT_MS": str(mode_01_wait_ms),
    "MODE_02_WAIT_MS": str(mode_02_wait_ms),
    "MODE_03_WAIT_MS": str(mode_03_wait_ms),
    "MODE_04_WAIT_MS": str(mode_04_wait_ms),
}

with st.expander("Notification preview", expanded=False):
    st.write("알림 설정:")
    st.write(f"- 데스크톱 알림: {'ON' if enable_desktop_alert else 'OFF'}")
    st.write(f"- 완료 알림: {'ON' if notify_on_success else 'OFF'}")
    st.write(f"- 에러 알림: {'ON' if notify_on_error else 'OFF'}")
    st.write(f"- 검증/중단 알림: {'ON' if notify_on_stop else 'OFF'}")
    st.write(f"- 영상 업로드 타임아웃 알림: {'ON' if notify_on_video_upload_timeout else 'OFF'}")

if campaign_mode == "BLOG_MIXED":
    st.subheader("BLOG_MIXED")
    st.caption("Mixed blog mode: the final creative in each adset is video, and all previous creatives are images.")
    blog_creative_count = st.number_input(
        "Ad creative count",
        min_value=1,
        max_value=100,
        value=int(env.get("AD_CREATIVE_COUNT", "4") or "4"),
        help="광고 소재 복제 수입니다. 실제 매체 소재 수는 입력값 + 1개입니다. 마지막 1개는 영상이고, 나머지는 이미지입니다. 예: 4 입력 -> 실제 5개, 이미지 4개 + 영상 1개.",
    )
    blog_actual_creative_count = int(blog_creative_count) + 1
    blog_image_count = max(blog_actual_creative_count - 1, 1)
    next_env["AD_CREATIVE_COUNT"] = str(blog_creative_count)
    blog_root = st.text_input("Blog asset root", value=env.get("BLOG_ASSET_ROOT", default_blog_root()))
    next_env["BLOG_ASSET_ROOT"] = blog_root

    with st.expander("Expected folder names", expanded=True):
        for index in range(1, int(adset_count) + 1):
            st.write(f"{index}. `{expected_blog_folder_name(index, daily_budget, schedule_time, blog_image_count)}`")

    st.subheader("Landing URLs")
    st.caption("VIDEO_ONLY_CBO preview처럼 확인하되, URL은 광고별이 아니라 광고세트 1개당 1개만 입력합니다.")
    blog_preview_rows = []
    for index in range(1, int(adset_count) + 1):
        key = f"BLOG_LANDING_URL_{index}"
        landing_url = st.text_input(f"Adset {index} landing URL", value=env.get(key, ""))
        next_env[key] = landing_url
        first_ad_index = ((index - 1) * blog_actual_creative_count) + 1
        image_names = [build_blog_image_ad_name(first_ad_index + offset) for offset in range(blog_image_count)]
        video_name = build_blog_video_ad_name(first_ad_index + blog_image_count)
        blog_preview_rows.append(
            {
                "adset_index": index,
                "adset_name": build_blog_adset_name(index),
                "landing_url": landing_url or "(missing)",
                "image_ads": ", ".join(image_names),
                "video_ad": video_name,
                "expected_folder": expected_blog_folder_name(index, daily_budget, schedule_time, blog_image_count),
            }
        )

    with st.expander("BLOG_MIXED preview / validation", expanded=True):
        st.write(f"Campaign name: `{campaign_name or '(missing)'}`")
        st.write(f"Adset count: `{adset_count}`")
        st.write(f"Actual creatives per adset: `{blog_actual_creative_count}`")
        st.write(f"Image creatives per adset: `{blog_image_count}`")
        st.write("Video creatives per adset: `1`")
        st.write(f"Blog asset root: `{blog_root}`")
        st.dataframe(blog_preview_rows, use_container_width=True)
        missing_urls = [row["adset_index"] for row in blog_preview_rows if row["landing_url"] == "(missing)"]
        if missing_urls:
            st.error(f"Landing URL missing for adset(s): {', '.join(map(str, missing_urls))}")
        else:
            st.success("All adset landing URLs are ready.")
elif campaign_mode in {"VIDEO_ONLY_CBO", "IMAGE_ONLY_CBO"}:
    is_image_cbo = campaign_mode == "IMAGE_ONLY_CBO"
    mode_prefix = "IMAGE_ONLY_CBO" if is_image_cbo else "VIDEO_ONLY_CBO"
    media_label = "Image" if is_image_cbo else "Video"
    media_label_lower = "image" if is_image_cbo else "video"
    folder_key = "IMAGE_ONLY_CBO_IMAGE_FOLDER" if is_image_cbo else "VIDEO_ONLY_CBO_VIDEO_FOLDER"
    st.subheader(campaign_mode)
    st.caption(f"{media_label}-only CBO campaign creation mode.")
    campaign_budget = st.text_input("Campaign budget", value=env.get("CAMPAIGN_BUDGET", "25000"))
    landing_path_number = st.number_input(
        "Repurely path number",
        min_value=1,
        max_value=999999,
        value=int(env.get("LANDING_PATH_NUMBER", env.get("REPURELY_PATH_NUMBER", "100")) or "100"),
        help="랜딩 URL의 /surl/P/{숫자} 부분입니다. 예: 99 입력 -> https://repurely.com/surl/P/99?...",
    )
    media_count = st.number_input(
        f"{media_label} ad count",
        min_value=1,
        max_value=100,
        value=int(env.get("AD_CREATIVE_COUNT", "0") or "0") + 1,
        help=f"입력한 숫자 그대로 실제 {media_label_lower} 광고 개수입니다. CBO 모드는 광고세트 1개 고정이며, 저장 시 .env의 AD_CREATIVE_COUNT에는 입력값 - 1로 저장됩니다. 예: 5 입력 -> 실제 광고 5개, AD_CREATIVE_COUNT=4",
    )
    media_folder = st.text_input(f"{media_label} file folder", value=env.get(folder_key, "./assets/images" if is_image_cbo else "./assets/videos"))
    next_env["CAMPAIGN_BUDGET"] = campaign_budget
    next_env["LANDING_PATH_NUMBER"] = str(landing_path_number)
    next_env["ADSET_COUNT"] = "0"
    next_env["AD_CREATIVE_COUNT"] = str(int(media_count) - 1)
    next_env[folder_key] = media_folder

    adset_total = 1
    creative_total = int(media_count)
    total_ads = adset_total * creative_total
    media_stems = list_image_stems_for_preview(media_folder) if is_image_cbo else list_video_stems_for_preview(media_folder)
    st.info(f"Budget preview for Meta: {format_budget_preview(campaign_budget) or 'invalid budget'}")

    preview_rows = []
    missing_count = 0
    st.subheader("Adset name")
    adset_names: dict[int, str] = {}
    for adset_index in range(1, adset_total + 1):
        key = f"{mode_prefix}_ADSET_NAME_{adset_index}"
        default_adset_name = build_image_only_cbo_adset_name(adset_index) if is_image_cbo else build_video_only_cbo_adset_name(adset_index)
        adset_name = st.text_input(f"Adset {adset_index} name", value=env.get(key, default_adset_name), key=key)
        next_env[key] = adset_name
        adset_names[adset_index] = adset_name

    st.subheader(f"Ads, {media_label_lower}s, and landing URL examples")
    for ad_index in range(1, total_ads + 1):
        adset_index = ((ad_index - 1) // creative_total) + 1
        ad_name_key = f"{mode_prefix}_AD_NAME_{ad_index}"
        default_builder = build_image_only_cbo_ad_name if is_image_cbo else build_video_only_cbo_ad_name
        default_ad_name = env.get(ad_name_key) or (media_stems[ad_index - 1] if ad_index <= len(media_stems) else default_builder(ad_index))
        ad_name = st.text_input(f"Ad {ad_index} name", value=default_ad_name, key=ad_name_key)
        next_env[ad_name_key] = ad_name
        media_file, expected = (find_image_file_for_preview(ad_name, media_folder) if is_image_cbo else find_video_file_for_preview(ad_name, media_folder))
        landing_url_example = default_landing_url(ad_name, str(landing_path_number))
        if not media_file:
            missing_count += 1
        preview_rows.append(
            {
                "adset": adset_names[adset_index],
                "ad_name": ad_name,
                f"{media_label_lower}_file": media_file or f"Missing: {', '.join(expected)}",
                "landing_url_example": landing_url_example,
            }
        )

    with st.expander(f"{campaign_mode} preview / validation", expanded=True):
        st.write(f"Campaign name: `{campaign_name or '(missing)'}`")
        st.write(f"Campaign budget: `{format_budget_preview(campaign_budget) or '(invalid)'}`")
        st.write(f"{media_label} folder: `{media_folder}`")
        st.write(f"Landing URL pattern: `{default_landing_url('{ad_name}', str(landing_path_number))}`")
        st.dataframe(preview_rows, use_container_width=True)
        if missing_count:
            st.error(f"{missing_count} ad rows have a missing {media_label_lower} file.")
        else:
            st.success(f"All {media_label_lower} files are ready. Landing URLs will be generated automatically.")
elif campaign_mode == "VIDEO_ONLY":
    st.subheader("VIDEO_ONLY")
    st.caption("Video-only direct landing mode.")
    landing_path_number = st.number_input(
        "Repurely path number",
        min_value=1,
        max_value=999999,
        value=int(env.get("LANDING_PATH_NUMBER", env.get("REPURELY_PATH_NUMBER", "100")) or "100"),
        help="랜딩 URL의 /surl/P/{숫자} 부분입니다. utm_campaign 값은 광고명으로 유지됩니다.",
    )
    video_count = st.number_input(
        "Video ad count",
        min_value=1,
        max_value=100,
        value=int(env.get("AD_CREATIVE_COUNT", "1") or "1"),
        help="기존 VIDEO_ONLY 흐름에서는 이 값이 복제 기준으로 저장되어 실제 영상 광고는 입력값 + 1개로 구성됩니다. 예: 1 입력 -> 실제 2개.",
    )
    next_env["LANDING_PATH_NUMBER"] = str(landing_path_number)
    video_root = st.text_input("Video asset root", value=env.get("VIDEO_ONLY_ASSET_ROOT", default_video_root()))
    next_env["AD_CREATIVE_COUNT"] = str(video_count)
    next_env["VIDEO_ONLY_ASSET_ROOT"] = video_root

    with st.expander("Expected folder names", expanded=True):
        detected_folder_names = read_child_folder_names(str(Path(video_root).parent), 1)
        matched = detected_folder_names
        if matched:
            st.caption("Detected from parent folder.")
            st.write(f"1. `{matched[0]}`")
        else:
            st.caption("No matching folder detected yet. Showing recommended folder name.")
            st.write(f"1. `{expected_video_folder_name()}`")
        st.caption("Videos may be directly inside that folder, or inside this child folder.")
        st.write(f"- `{expected_video_child_folder_name()}`")

    with st.expander("Landing URL pattern", expanded=True):
        mmdd = datetime.now().strftime("%m%d")
        st.write(f"`{default_landing_url(f'f_v_o_l_{mmdd}_1', str(landing_path_number))}`")
        st.write(f"`{default_landing_url(f'f_v_o_l_{mmdd}_2', str(landing_path_number))}`")
else:
    st.subheader("IMAGE_ONLY")
    landing_path_number = st.number_input(
        "Repurely path number",
        min_value=1,
        max_value=999999,
        value=int(env.get("LANDING_PATH_NUMBER", env.get("REPURELY_PATH_NUMBER", "100")) or "100"),
        help="랜딩 URL의 /surl/P/{숫자} 부분입니다. utm_campaign 값은 광고명으로 유지됩니다.",
    )
    creative_count = st.number_input(
        "Image ad count",
        min_value=1,
        max_value=100,
        value=int(env.get("AD_CREATIVE_COUNT", "4") or "4"),
        help="기존 IMAGE_ONLY 흐름에서는 이 값이 복제 기준으로 저장되어 실제 이미지 광고는 입력값 + 1개로 구성됩니다. 예: 4 입력 -> 실제 5개.",
    )
    next_env["LANDING_PATH_NUMBER"] = str(landing_path_number)
    per_ad_upload = st.checkbox(
        "Upload one image per ad",
        value=env.get("IMAGE_ONLY_UPLOAD_MODE", "PER_AD").upper() != "LEGACY",
    )
    media_folder = st.text_input(
        "Image media folder",
        value=env.get("IMAGE_ONLY_ASSET_ROOT") or env.get("MEDIA_FOLDER_PATH") or default_image_root(),
    )
    image_effective_adset_count = int(adset_count) + 1
    with st.expander("Expected folder names", expanded=True):
        detected_folder_names = read_child_folder_names(media_folder, image_effective_adset_count)
        if detected_folder_names:
            st.caption("Detected from Image media folder.")
        else:
            st.caption("No child folders detected yet. Showing recommended folder names.")
        for index in range(1, image_effective_adset_count + 1):
            folder_name = detected_folder_names[index - 1] if index <= len(detected_folder_names) else expected_image_folder_name(index, daily_budget, int(creative_count) + 1, schedule_time)
            st.write(f"{index}. `{folder_name}`")
    with st.expander("Landing URL examples", expanded=True):
        actual_creative_count = int(creative_count) + 1
        total_image_ads = image_effective_adset_count * actual_creative_count
        st.caption("실제 자동화에서는 직접 입력 없이 광고명 기준으로 자동 생성됩니다.")
        st.write(f"Landing URL pattern: `{default_landing_url('{ad_name}', str(landing_path_number))}`")
        st.write(f"Total generated image URLs: `{total_image_ads}`")
        for index in range(1, min(total_image_ads, 5) + 1):
            ad_name = build_image_only_ad_name(index)
            st.write(f"{index}. `{default_landing_url(ad_name, str(landing_path_number))}`")
        if total_image_ads > 5:
            st.caption(f"... plus {total_image_ads - 5} more")
    next_env["AD_CREATIVE_COUNT"] = str(creative_count)
    next_env["IMAGE_ONLY_UPLOAD_MODE"] = "PER_AD" if per_ad_upload else "LEGACY"
    next_env["IMAGE_ONLY_ASSET_ROOT"] = media_folder if per_ad_upload else ""
    next_env["MEDIA_FOLDER_PATH"] = media_folder

st.divider()

col1, col2, col3 = st.columns(3)
with col1:
    if st.button("Save .env", type="primary"):
        saved = write_env(next_env)
        st.success(f"Saved: {ENV_PATH}")
        st.toast("알림 설정을 포함해 .env 저장 완료")
        st.code(saved, language="dotenv")

with col2:
    if st.button("Run dry-run"):
        dry_run_env = {**next_env, "DRY_RUN": "true"}
        errors = validate_form(dry_run_env)
        saved = write_env(dry_run_env)
        st.caption("Saved .env for dry-run")
        st.code(saved, language="dotenv")
        if errors:
            st.error("Fix these fields before dry-run:")
            st.toast("작업 중단: 입력값 검증 실패")
            for error in errors:
                st.write(f"- {error}")
        else:
            command = ["node", "src/open-campaign.js"]
            code, output = run_command(command, timeout=300)
            show_command_result("dry-run", command, code, output)
            if code == 0:
                st.toast("Meta 광고 자동화 dry-run 완료")
            else:
                st.toast("작업 중단: 자세한 내용은 로그를 확인하세요")

with col3:
    if st.button("Run real automation terminal", disabled=not IS_WINDOWS):
        real_env = {**next_env, "DRY_RUN": "false"}
        errors = validate_form(real_env)
        saved = write_env(real_env)
        st.caption("Saved .env for real run")
        st.code(saved, language="dotenv")
        if errors:
            st.error("Fix these fields before real run:")
            st.toast("작업 중단: 입력값 검증 실패")
            for error in errors:
                st.write(f"- {error}")
        else:
            try:
                start_powershell("npm run open-campaign")
                st.warning("DRY_RUN=false saved. Automation terminal opened.")
                st.toast("자동화 터미널을 열었습니다. 완료/에러 알림은 실제 실행 결과에 따라 표시됩니다")
            except Exception as exc:
                st.error(f"Automation terminal launcher is unavailable here: {exc}")
                st.info("Streamlit Cloud can preview settings, but real Meta automation must run on your local Windows PC.")

with st.expander("Current .env", expanded=False):
    if ENV_PATH.exists():
        st.code(ENV_PATH.read_text(encoding="utf-8"), language="dotenv")
    else:
        st.info(".env does not exist yet.")
